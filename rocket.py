import re
from openpyxl import load_workbook

#calc change in mass using old pressure
def massOldPressure():
    return densityWater * (volumeRocket - (initalpressureTotal * volumeAir ** 1.4 / (pressureTotal - 1.4 * (pressureTotal ** (12/7) * areaExhaust * (2 * (pressureTotal - pressureAtmosphere) / densityWater) ** 0.5 * timeInterval / (initalpressureTotal ** (5/7) * volumeAir)))) ** (1/1.4))

#calc new pressure
def newPressure():
    return pressureTotal - 1.4 * (pressureTotal ** (12/7) * areaExhaust * (2*(pressureTotal - pressureAtmosphere) / densityWater) ** 0.5 * timeInterval / (initalpressureTotal ** (5/7) * volumeAir))

#calc change in mass using new pressure
def massChange():
    return densityWater * (volumeRocket - (initalpressureTotal * volumeAir ** 1.4 / pressureTotal) ** (1/1.4))

#write to spreadsheet
def write(r, worksheet):
    worksheet.cell(row=r, column=1).value = time
    worksheet.cell(row=r, column=2).value = pressureTotal
    worksheet.cell(row=r, column=3).value = forceThrust
    worksheet.cell(row=r, column=4).value = forceDrag
    worksheet.cell(row=r, column=5).value = massTotal
    worksheet.cell(row=r, column=6).value = acceleration
    worksheet.cell(row=r, column=7).value = velocity
    worksheet.cell(row=r, column=8).value = height

#read parameter file, convert to numbers
rawParameters = open("parameters.txt", "r")
parameters = rawParameters.readlines()

for x in range(len(parameters)):
    parameters[x] = re.sub("[^\d\.]", "", parameters[x])
    parameters[x] = parameters[x].strip()

rawParameters.close()

#initial parameter variables
massRocket = float(parameters[0])
volumeRocket = float(parameters[1])
areaExhaust = float(parameters[2])
pressureGauge = float(parameters[3])
pressureAtmosphere = float(parameters[4])
dragCoeff = float(parameters[5])
areaCrossSection = float(parameters[6])
massWater = float(parameters[7])

#initial non parameter variables and constants
initalpressureTotal = pressureGauge + pressureAtmosphere
pressureTotal = initalpressureTotal
densityWater = 1000
accelerationG = -9.81
densityAir = 1.2
volumeWater = massWater / densityWater
volumeAir = volumeRocket - volumeWater
time = 0
timeInterval = 0.01

#initial dynamics variables
forceThrust = 2 * (pressureTotal - pressureAtmosphere) * areaExhaust
forceDrag = 0
massTotal = massRocket + massWater
acceleration = (forceThrust + forceDrag) / massTotal + accelerationG
velocity = 0
height = 0
tempAcceleration = 0
tempVelocity = 0




#create new sheet
wb = load_workbook("flight_model.xlsx")
wb.create_sheet("Flight", 0)
ws = wb.active

ws.cell(row=1, column=1).value = "Time"
ws.cell(row=1, column=2).value = "Total Pressure"
ws.cell(row=1, column=3).value = "Thrust"
ws.cell(row=1, column=4).value = "Drag"
ws.cell(row=1, column=5).value = "Mass"
ws.cell(row=1, column=6).value = "Acceleration"
ws.cell(row=1, column=7).value = "Velocity"
ws.cell(row=1, column=8).value = "Position"

row = 2


while(height >= 0):
    write(row, ws)
    row += 1
    time += timeInterval
    #total pressure calc
    if massOldPressure() > 0:
        pressureTotal = newPressure()
    else:
        pressureTotal = pressureAtmosphere
    #forces calc
    forceThrust = 2 * (pressureTotal - pressureAtmosphere) * areaExhaust
    forceDrag = -0.5 * dragCoeff * areaCrossSection * densityAir * velocity * abs(velocity)
    #mass calc
    if massChange() > 0:
        massTotal = massChange() + massRocket
    else:
        massTotal = massRocket
    #kinematics calc
    tempAcceleration = acceleration
    acceleration = (forceThrust + forceDrag) / massTotal + accelerationG
    tempVelocity = velocity
    velocity += tempAcceleration * timeInterval
    height += (tempVelocity + velocity) / 2 * timeInterval

write(row, ws)

wb.save("flight_model.xlsx")


